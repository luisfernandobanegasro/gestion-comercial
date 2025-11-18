# reportes/services.py
from io import BytesIO
from typing import Tuple, Dict, Any, List, Optional

from datetime import timedelta
from django.utils import timezone
from django.db.models import Sum, F, Avg, Q
from django.db.models.functions import TruncDay, TruncMonth

from ventas.models import Venta, ItemVenta
from clientes.models import Cliente
from catalogo.models import Producto

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


# =========================
# Consultas de ventas
# =========================

def consultar_ventas(
    start_date: Optional[str],
    end_date: Optional[str],
    group_by: str,
    cliente: Optional[str] = None,
    contiene: Optional[str] = None,
) -> Tuple[List[str], List[List]]:
    """
    Reportes de ventas.

    - group_by = 'producto'  → productos vendidos
    - group_by = 'cliente'   → ventas por cliente
    - group_by = 'categoria' → ventas por categoría
    """
    qs = ItemVenta.objects.select_related(
        "venta", "producto", "venta__cliente", "producto__categoria"
    )

    if start_date:
        qs = qs.filter(venta__creado_en__date__gte=start_date)
    if end_date:
        qs = qs.filter(venta__creado_en__date__lte=end_date)

    qs = qs.filter(venta__estado__in=["pagada"])

    # Filtro por cliente (nombre o documento)
    if cliente:
        qs = qs.filter(
            Q(venta__cliente__nombre__icontains=cliente)
            | Q(venta__cliente__documento__icontains=cliente)
        )

    # Filtro por producto que "contiene"
    if contiene:
        qs = qs.filter(producto__nombre__icontains=contiene)

    rows: List[List] = []
    headers: List[str] = []

    if group_by == "producto":
        agg = (
            qs.values(nombre=F("producto__nombre"))
            .annotate(
                cantidad=Sum("cantidad"),
                monto=Sum(F("subtotal")),
            )
            .order_by("-monto", "-cantidad", "nombre")
        )
        headers = ["Producto", "Cantidad total", "Monto total"]
        rows = [
            [r["nombre"], int(r["cantidad"] or 0), float(r["monto"] or 0.0)]
            for r in agg
        ]

    elif group_by == "cliente":
        agg = (
            qs.values(
                cliente=F("venta__cliente__nombre"),
                doc=F("venta__cliente__documento"),
            )
            .annotate(
                cantidad=Sum("cantidad"),
                monto=Sum(F("subtotal")),
            )
            .order_by("-monto", "-cantidad", "cliente")
        )
        headers = ["Cliente", "Documento", "Cantidad total", "Monto total"]
        rows = [
            [
                r["cliente"] or "SIN NOMBRE",
                r["doc"] or "",
                int(r["cantidad"] or 0),
                float(r["monto"] or 0.0),
            ]
            for r in agg
        ]

    elif group_by == "categoria":
        agg = (
            qs.values(cat=F("producto__categoria__nombre"))
            .annotate(
                cantidad=Sum("cantidad"),
                monto=Sum(F("subtotal")),
            )
            .order_by("-monto", "-cantidad", "cat")
        )
        headers = ["Categoría", "Cantidad total", "Monto total"]
        rows = [
            [
                r["cat"] or "SIN CATEGORÍA",
                int(r["cantidad"] or 0),
                float(r["monto"] or 0.0),
            ]
            for r in agg
        ]
    else:
        # Detalle por venta (folio)
        agg = (
            qs.values(
                folio=F("venta__folio"),
                fecha=F("venta__creado_en__date"),
                cliente=F("venta__cliente__nombre"),
            )
            .annotate(
                cantidad=Sum("cantidad"),
                monto=Sum(F("subtotal")),
            )
            .order_by("-fecha", "-monto")
        )
        headers = ["Folio", "Fecha", "Cliente", "Cantidad total", "Monto total"]
        rows = [
            [
                r["folio"],
                r["fecha"].strftime("%Y-%m-%d") if r["fecha"] else "",
                r["cliente"] or "SIN NOMBRE",
                int(r["cantidad"] or 0),
                float(r["monto"] or 0.0),
            ]
            for r in agg
        ]

    return headers, rows


def consultar_top_productos(
    start_date: Optional[str],
    end_date: Optional[str],
    limit: Optional[int],
    categoria: Optional[str] = None,
) -> Tuple[List[str], List[List]]:
    """Ranking de productos más vendidos en un período."""
    qs = ItemVenta.objects.select_related("venta", "producto", "producto__categoria")
    if start_date:
        qs = qs.filter(venta__creado_en__date__gte=start_date)
    if end_date:
        qs = qs.filter(venta__creado_en__date__lte=end_date)
    qs = qs.filter(venta__estado__in=["pagada"])
    if categoria:
        qs = qs.filter(producto__categoria__nombre__icontains=categoria)

    agg = (
        qs.values(n=F("producto__nombre"))
        .annotate(cantidad=Sum("cantidad"), monto=Sum(F("subtotal")))
        .order_by("-cantidad", "-monto", "n")
    )
    if limit:
        agg = agg[:limit]

    headers = ["Producto", "Cantidad vendida", "Monto total"]
    rows = [
        [r["n"], int(r["cantidad"] or 0), float(r["monto"] or 0)] for r in agg
    ]
    return headers, rows


def consultar_sin_movimiento(
    start_date: Optional[str],
    end_date: Optional[str],
    categoria: Optional[str] = None,
) -> Tuple[List[str], List[List]]:
    """Productos que NO tuvieron ventas en el período."""
    vendibles = ItemVenta.objects.select_related("venta", "producto")
    if start_date:
        vendibles = vendibles.filter(venta__creado_en__date__gte=start_date)
    if end_date:
        vendibles = vendibles.filter(venta__creado_en__date__lte=end_date)
    vendibles = vendibles.values_list("producto_id", flat=True).distinct()

    qs = Producto.objects.exclude(id__in=vendibles).select_related("categoria")
    if categoria:
        qs = qs.filter(categoria__nombre__icontains=categoria)

    headers = ["Producto", "Categoría", "Stock", "Precio", "Observación"]
    rows = []
    for p in qs.order_by("nombre"):
        rows.append(
            [
                p.nombre,
                getattr(p.categoria, "nombre", "")
                if getattr(p, "categoria", None)
                else "",
                getattr(p, "stock", 0),
                float(getattr(p, "precio", 0.0)),
                "Sin ventas en el período",
            ]
        )
    return headers, rows


# =========================
# Stock / Precios
# =========================

def consultar_stock(
    categoria: Optional[str] = None,
    marca: Optional[str] = None,
    contiene: Optional[str] = None,
):
    qs = Producto.objects.select_related("categoria")
    if categoria:
        qs = qs.filter(categoria__nombre__icontains=categoria)
    if marca and hasattr(Producto, "marca"):
        qs = qs.filter(marca__nombre__icontains=marca)
    if contiene:
        qs = qs.filter(nombre__icontains=contiene)

    headers = ["Producto", "Categoría", "Stock", "Stock mínimo", "Precio"]
    rows = []
    for p in qs.order_by("nombre"):
        rows.append(
            [
                p.nombre,
                getattr(p.categoria, "nombre", "")
                if getattr(p, "categoria", None)
                else "",
                getattr(p, "stock", 0),
                getattr(p, "stock_minimo", 0),
                float(getattr(p, "precio", 0.0)),
            ]
        )
    return headers, rows


def consultar_stock_bajo(
    threshold: Optional[int] = None,
    categoria: Optional[str] = None,
    limit: Optional[int] = None,
):
    qs = Producto.objects.select_related("categoria")

    if categoria:
        qs = qs.filter(categoria__nombre__icontains=categoria)

    if threshold is not None:
        qs = qs.filter(stock__lt=threshold)
    elif hasattr(Producto, "stock_minimo"):
        qs = qs.filter(stock__lt=F("stock_minimo"))
    else:
        qs = qs.filter(stock__lt=10)

    qs = qs.order_by("stock", "nombre")
    if limit:
        qs = qs[:limit]

    headers = ["Producto", "Categoría", "Stock", "Stock mínimo", "Sugerencia"]
    rows = []
    for p in qs:
        minimo = getattr(p, "stock_minimo", None)
        sug = f"Reponer hasta {minimo}" if minimo else "Reponer"
        rows.append(
            [
                p.nombre,
                getattr(p.categoria, "nombre", "")
                if getattr(p, "categoria", None)
                else "",
                getattr(p, "stock", 0),
                minimo if minimo is not None else "-",
                sug,
            ]
        )
    return headers, rows


def consultar_precios(
    categoria: Optional[str] = None,
    marca: Optional[str] = None,
    contiene: Optional[str] = None,
):
    qs = Producto.objects.select_related("categoria")
    if categoria:
        qs = qs.filter(categoria__nombre__icontains=categoria)
    if marca and hasattr(Producto, "marca"):
        qs = qs.filter(marca__nombre__icontains=marca)
    if contiene:
        qs = qs.filter(nombre__icontains=contiene)

    headers = ["Producto", "Categoría", "Precio", "Stock"]
    rows = []
    for p in qs.order_by("nombre"):
        rows.append(
            [
                p.nombre,
                getattr(p.categoria, "nombre", "")
                if getattr(p, "categoria", None)
                else "",
                float(getattr(p, "precio", 0.0)),
                getattr(p, "stock", 0),
            ]
        )
    return headers, rows


# =========================
# Exportadores Excel / PDF
# =========================

def generar_excel(headers: List[str], rows: List[List]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(headers)
    header_fill = PatternFill(
        start_color="1E293B", end_color="1E293B", fill_type="solid"
    )
    header_ink = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_ink
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(r)

    ws.freeze_panes = "A2"

    if headers and "monto" in headers[-1].lower():
        for row_idx in range(2, ws.max_row + 1):
            c = ws.cell(row=row_idx, column=len(headers))
            try:
                float(c.value)
                c.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                c.alignment = Alignment(horizontal="right")
            except Exception:
                pass

    for idx, h in enumerate(headers, start=1):
        if "cantidad" in h.lower():
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=idx).alignment = Alignment(
                    horizontal="center"
                )

    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border = border

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(12, max_len + 2), 42
        )

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generar_pdf(headers: List[str], rows: List[List], meta: Dict[str, Any]) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="Title",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=10,
    )
    meta_style = ParagraphStyle(
        name="Meta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=8,
    )

    elems = []
    elems.append(Paragraph("Reporte", title_style))

    rango = f"Rango: {meta.get('start_date') or '-'} a {meta.get('end_date') or '-'}"
    detalles = []
    if meta.get("group_by"):
        detalles.append(f"Grupo: {meta.get('group_by')}")
    if meta.get("intent"):
        detalles.append(f"Intento: {meta.get('intent')}")
    if meta.get("metrics"):
        detalles.append(f"Métricas: {', '.join(meta.get('metrics'))}")
    sub = rango + (f" | {' | '.join(detalles)}" if detalles else "")
    elems.append(Paragraph(sub, meta_style))
    elems.append(Spacer(1, 6))

    data = []
    if headers:
        data.append(headers)
    data.extend(rows or [])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f8fafc")],
                ),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    if headers and any(k in headers[-1].lower() for k in ("monto", "total", "precio")):
        table.setStyle(TableStyle([("ALIGN", (-1, 1), (-1, -1), "RIGHT")]))

    elems.append(table)
    doc.build(elems)
    pdf = buf.getvalue()
    buf.close()
    return pdf


# =========================
# Datos para Dashboard
# =========================

def get_dashboard_data() -> Dict[str, Any]:
    now = timezone.now()
    today = now.date()
    start_of_month = today.replace(day=1)
    start_of_30_days = today - timedelta(days=30)

    ventas_pagadas = Venta.objects.filter(estado="pagada")

    ventas_hoy = (
        ventas_pagadas.filter(creado_en__date=today)
        .aggregate(total=Sum("total"))["total"]
        or 0
    )
    ventas_mes = (
        ventas_pagadas.filter(creado_en__gte=start_of_month)
        .aggregate(total=Sum("total"))["total"]
        or 0
    )
    ticket_promedio = (
        ventas_pagadas.filter(creado_en__gte=start_of_month)
        .aggregate(avg=Avg("total"))["avg"]
        or 0
    )
    nuevos_clientes_mes = Cliente.objects.filter(
        creado_en__gte=start_of_month
    ).count()

    ventas_diarias = (
        ventas_pagadas.filter(creado_en__gte=start_of_30_days)
        .annotate(dia=TruncDay("creado_en"))
        .values("dia")
        .annotate(total=Sum("total"))
        .order_by("dia")
    )

    ventas_por_categoria = (
        ventas_pagadas.filter(creado_en__gte=start_of_month)
        .values("items__producto__categoria__nombre")
        .annotate(total=Sum("items__subtotal"))
        .order_by("-total")
    )

    ultimas_ventas = ventas_pagadas.order_by("-creado_en")[:5].select_related("cliente")

    return {
        "kpis": {
            "ventas_hoy": f"{ventas_hoy:,.2f}",
            "ventas_mes_actual": f"{ventas_mes:,.2f}",
            "nuevos_clientes_mes": nuevos_clientes_mes,
            "ticket_promedio": f"{ticket_promedio:,.2f}",
        },
        "ventas_ultimos_30_dias": [
            {"fecha": item["dia"].strftime("%Y-%m-%d"), "total": float(item["total"])}
            for item in ventas_diarias
        ],
        "ventas_por_categoria": [
            {
                "categoria": item["items__producto__categoria__nombre"],
                "total": float(item["total"]),
            }
            for item in ventas_por_categoria
            if item["items__producto__categoria__nombre"]
        ],
        "ultimas_ventas": [
            {
                "id": v.id,
                "folio": v.folio,
                "cliente": v.cliente.nombre if v.cliente else "N/A",
                "total": f"{v.total:,.2f}",
                "fecha": v.creado_en.strftime("%Y-%m-%d %H:%M"),
            }
            for v in ultimas_ventas
        ],
    }
