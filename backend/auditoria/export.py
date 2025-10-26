from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io

def exportar_auditoria_excel(queryset):
    """Genera un archivo Excel a partir de un queryset de registros de auditoría."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="auditoria.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"

    # Encabezados
    headers = ["Fecha y Hora", "Usuario", "Módulo", "Acción", "Ruta", "Método", "Estado", "IP"]
    ws.append(headers)

    # Estilo para encabezados
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for registro in queryset:
        ws.append([
            registro.creado_en.strftime("%Y-%m-%d %H:%M:%S"),
            registro.usuario.username if registro.usuario else "Anónimo",
            registro.modulo,
            registro.accion,
            registro.ruta,
            registro.metodo,
            registro.estado,
            registro.ip,
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

def exportar_auditoria_pdf(queryset):
    """Genera un archivo PDF a partir de un queryset de registros de auditoría."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph("Bitácora de Auditoría", styles['h1']))

    # Datos de la tabla
    data = [["Fecha/Hora", "Usuario", "Módulo", "Acción", "Ruta", "Método", "Estado", "IP"]]
    for registro in queryset:
        data.append([
            registro.creado_en.strftime("%y-%m-%d %H:%M"),
            registro.usuario.username if registro.usuario else "Anónimo",
            registro.modulo, registro.accion, registro.ruta, registro.metodo, registro.estado, registro.ip
        ])

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])

    # Crear y añadir la tabla
    table = Table(data)
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="auditoria.pdf"'})
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import io

def exportar_auditoria_excel(queryset):
    """Genera un archivo Excel a partir de un queryset de registros de auditoría."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="auditoria.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría"

    # Encabezados
    headers = ["Fecha y Hora", "Usuario", "Módulo", "Acción", "Ruta", "Método", "Estado", "IP"]
    ws.append(headers)

    # Estilo para encabezados
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for registro in queryset:
        ws.append([
            registro.creado_en.strftime("%Y-%m-%d %H:%M:%S"),
            registro.usuario.username if registro.usuario else "Anónimo",
            registro.modulo,
            registro.accion,
            registro.ruta,
            registro.metodo,
            registro.estado,
            registro.ip,
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

def exportar_auditoria_pdf(queryset):
    """Genera un archivo PDF a partir de un queryset de registros de auditoría."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph("Bitácora de Auditoría", styles['h1']))

    # Datos de la tabla
    data = [["Fecha/Hora", "Usuario", "Módulo", "Acción", "Ruta", "Método", "Estado", "IP"]]
    for registro in queryset:
        data.append([
            registro.creado_en.strftime("%y-%m-%d %H:%M"),
            registro.usuario.username if registro.usuario else "Anónimo",
            registro.modulo, registro.accion, registro.ruta, registro.metodo, registro.estado, registro.ip
        ])

    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])

    # Crear y añadir la tabla
    table = Table(data)
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="auditoria.pdf"'})